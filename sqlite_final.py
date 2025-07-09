import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sqlite3

# === INPUT FILES ===  ## these are the results of ocr, defect detection and including rohs non compliant
reference_file = r"C:\Users\HARSHITHA\Downloads\Python\my\New Final_Traceability_Updated.csv"
ocr_file = r"C:\Users\HARSHITHA\Downloads\Python\my\total_label_with_final_result_with_rohs.csv"
hazardous_file = r"C:\Users\HARSHITHA\Downloads\Python\my\Rejected_Log.csv"
defect_detection_result_file = r"C:\Users\HARSHITHA\Downloads\Python\my\94d2652c-217c-46c6-b260-9f37a5ee8f17 (1).xlsx"
output_file = "final_combined_result5.xlsx"
db_path = "inspection_results.db"
table_name = "FinalInspectionResults"

# === File Loader ===
def load_file(filename):
    ext = os.path.splitext(filename)[-1].lower()
    if ext == ".csv":
        return pd.read_csv(filename)
    elif ext == ".xlsx":
        return pd.read_excel(filename)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

# === Load datasets ===
reference_df = load_file(reference_file)
ocr_df = load_file(ocr_file)
rohs_no_df = load_file(hazardous_file)
label_df = load_file(defect_detection_result_file)

# === Normalize and Rename columns ===
reference_df.rename(columns={'DeviceID': 'Device ID', 'BatchID': 'Batch ID'}, inplace=True)
ocr_df.rename(columns={'RoHSCompliance': 'RoHSCompliance', 'final_result': 'final_result'}, inplace=True)
label_df.rename(columns={'BatchID': 'Batch ID', 'True_Label': 'True_label', 'Predicted_Label': 'Predicted_label'}, inplace=True)

reference_df['RoHSCompliance'] = reference_df['RoHSCompliance'].astype(str).str.strip().str.lower()
ocr_df['RoHSCompliance'] = ocr_df['RoHSCompliance'].astype(str).str.strip().str.lower()
ocr_df['final_result'] = ocr_df['final_result'].astype(str).str.strip().str.upper()
label_df['Batch ID'] = label_df['Batch ID'].astype(str).str.strip()
label_df['True_label'] = label_df['True_label'].astype(str).str.strip().str.lower()
label_df['Predicted_label'] = label_df['Predicted_label'].astype(str).str.strip().str.lower()

# === Fill missing Device ID using reference ===
def fill_device_id(row):
    if pd.isna(row['Device ID']) or row['Device ID'] == "":
        ref_match = reference_df[reference_df['Batch ID'] == row['Batch ID']]
        if not ref_match.empty:
            return ref_match.iloc[0]['Device ID']
    return row['Device ID']

ocr_df['Device ID'] = ocr_df.apply(fill_device_id, axis=1)

# === Batch-Wise Result Creation ===
def batch_result(batch_group):
    batch_id = batch_group['Batch ID'].iloc[0]
    device_id = batch_group['Device ID'].iloc[0]
    rohs = batch_group['RoHSCompliance'].iloc[0]

    timestamp = batch_group['Timestamp'].iloc[0] if 'Timestamp' in batch_group.columns else ""
    if pd.isna(timestamp) or str(timestamp).strip() == "":
        timestamp = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
    else:
        try:
            timestamp = datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S').strftime('%d-%m-%Y %H:%M:%S')
        except:
            try:
                timestamp = datetime.strptime(str(timestamp), '%d-%m-%Y %H:%M').strftime('%d-%m-%Y %H:%M:%S')
            except:
                timestamp = datetime.now().strftime('%d-%m-%Y %H:%M:%S')

    all_approved = (batch_group['final_result'] == "APPROVED").all()

    pred_row = label_df[label_df['Batch ID'] == batch_id]
    true_label = pred_row.iloc[0]['True_label'] if not pred_row.empty else ""
    pred_label = pred_row.iloc[0]['Predicted_label'] if not pred_row.empty else ""

    if rohs == "no":
        return pd.Series({
            'Device ID': device_id,
            'Batch ID': batch_id,
            'Timestamp': timestamp,
            'RoHSCompliance': rohs,
            'Inspection Result': 'REJECTED',
            'Failed Reason': 'Hazardous Substances Found'
        })

    if all_approved:
        if true_label == "good" and pred_label == "good":
            return pd.Series({
                'Device ID': device_id,
                'Batch ID': batch_id,
                'Timestamp': timestamp,
                'RoHSCompliance': rohs,
                'Inspection Result': 'APPROVED',
                'Failed Reason': ''
            })
        else:
            return pd.Series({
                'Device ID': device_id,
                'Batch ID': batch_id,
                'Timestamp': timestamp,
                'RoHSCompliance': rohs,
                'Inspection Result': 'REJECTED',
                'Failed Reason': 'Defective Product Detected'
            })
    else:
        return pd.Series({
            'Device ID': device_id,
            'Batch ID': batch_id,
            'Timestamp': timestamp,
            'RoHSCompliance': rohs,
            'Inspection Result': 'REJECTED',
            'Failed Reason': 'Blurry Label or Mismatch'
        })

# === Apply validation ===
final_df = ocr_df.groupby('Batch ID').apply(batch_result).reset_index(drop=True)

# === Add RoHS = no entries if not already present ===
rohs_no_df.rename(columns={'DeviceID': 'Device ID', 'BatchID': 'Batch ID'}, inplace=True)
rohs_no_df['RoHSCompliance'] = rohs_no_df['RoHSCompliance'].astype(str).str.strip().str.lower()
rohs_no_df = rohs_no_df[rohs_no_df['RoHSCompliance'] == 'no']

if 'Timestamp' not in rohs_no_df.columns:
    rohs_no_df['Timestamp'] = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
else:
    rohs_no_df['Timestamp'] = rohs_no_df['Timestamp'].fillna('').astype(str)
    rohs_no_df['Timestamp'] = rohs_no_df['Timestamp'].apply(
        lambda x: datetime.now().strftime('%d-%m-%Y %H:%M:%S') if x.strip() == '' else x
    )

rohs_no_df = rohs_no_df[['Device ID', 'Batch ID', 'Timestamp', 'RoHSCompliance']].drop_duplicates()
rohs_no_df['Inspection Result'] = 'REJECTED'
rohs_no_df['Failed Reason'] = 'Hazardous Substances Found'

new_rohs = rohs_no_df[~rohs_no_df['Batch ID'].isin(final_df['Batch ID'])]
final_df = pd.concat([final_df, new_rohs], ignore_index=True)

# === Save to Excel ===
final_df.to_excel(output_file, index=False)

# === Apply Color Coding ===
wb = load_workbook(output_file)
ws = wb.active

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

header = [cell.value for cell in ws[1]]
result_col_idx = header.index('Inspection Result') + 1

for row in range(2, ws.max_row + 1):
    result = ws.cell(row=row, column=result_col_idx).value
    fill = green_fill if result == "APPROVED" else red_fill
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = fill

wb.save(output_file)
print(f"✅ Final Excel saved: {output_file} with results, timestamps in 'dd-mm-yyyy HH:MM:SS', and conditional formatting.")

# === Save Final Results to SQLite DB ===
conn = sqlite3.connect(db_path)
final_df.to_sql(table_name, conn, if_exists='replace', index=False)
print(f"✅ Final results also saved to SQLite database: {db_path}, table: {table_name}")

# === Optional Preview ===
preview = pd.read_sql_query(f"SELECT * FROM {table_name} LIMIT 5", conn)
print("\n📋 Preview of FinalInspectionResults:")
print(preview)

conn.close()
